from openpyxl import load_workbook
import os
import cv2
import shutil


def copy_images_to_new_path(img_paths, dest_dir):
    """
    将多个图片复制到新的目录中
    :param img_paths: 图片路径列表
    :param dest_dir: 目标目录路径
    """
    # 如果目标目录不存在，先创建
    os.makedirs(dest_dir, exist_ok=True)

    for img_path in img_paths:
        if os.path.exists(img_path):
            # 取出文件名
            file_name = os.path.basename(img_path)
            # 拼接目标路径
            dest_path = os.path.join(dest_dir, file_name)
            # 复制文件
            shutil.copy(img_path, dest_path)
            print(f"已复制: {img_path} → {dest_path}")
        else:
            print(f"文件不存在: {img_path}")


#  NO! 216_e_0 & 216_i_1 & 144_k_289
# 加载工作簿
workbook = load_workbook(filename=r'E:\Xyj\812_CLIP\Generate_DFME\testB_full_with_AU.xlsx')
sheet = workbook['Sheet1']

# 读取整个工作表数据
xlsx_values = []
for row in sheet.iter_rows(min_row=1, max_col=8, values_only=True):
    xlsx_values.append(row)

del xlsx_values[0]  # 删除标题行

# 情感列表
main_path = r"E:\Xyj\DFME\test_data_B\test_data_B"
base_path = r"E:\Xyj\812_CLIP\Generate_DFME\DFME_testb"
num = 0
for row in xlsx_values:
    Subject = row[0]
    Filename = row[1]
    OnsetFrame = str(row[2])
    OnsetFrame = OnsetFrame.zfill(5)

    ApexFrame = str(row[3])
    ApexFrame = ApexFrame.zfill(5)

    one = os.path.join(main_path, f"{Filename}", f"{OnsetFrame}.png")
    apex = os.path.join(main_path, f"{Filename}", f"{ApexFrame}.png")
    dst_image = os.path.join(base_path, f"{Filename}")
    copy_images_to_new_path([one, apex], dst_image)
    num += 1
print(num)
