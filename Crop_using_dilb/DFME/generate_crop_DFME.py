import os
import cv2
import dlib
from openpyxl import load_workbook
import shutil


# 加载 dlib 的人脸检测器和面部标志点预测器
detector = dlib.get_frontal_face_detector()
predictor = dlib.shape_predictor('shape_predictor_68_face_landmarks.dat')


def crop_face(image_path, output_path="cropped_face.jpg"):
    """
    从输入图片中识别最大的人脸并进行切割，去除前16个关键点后，调整剪裁宽度为两个眉毛最外侧点的宽度，并将矩形框的高度向上增加10像素。

    :param image_path: 输入的图片路径
    :param output_path: 切割后图片的保存路径，默认保存为 "cropped_face.jpg"
    :return: None
    """
    # 加载输入图像
    image = cv2.imread(image_path)

    # 如果图像加载失败，返回错误
    if image is None:
        print(f"图像加载失败，请检查图片路径：{image_path}")
        return

    # 转换为灰度图像
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # 检测人脸
    faces = detector(gray, 1)

    if len(faces) > 0:
        # 选择最大的人脸矩形框
        largest_face = max(faces, key=lambda rect: rect.width() * rect.height())

        # 获取面部关键点
        landmarks = predictor(gray, largest_face)

        # 获取眉毛最外侧点（第17, 18, 19, 20 对应左眉，21, 22, 23, 24 对应右眉）
        left_brow_left = landmarks.part(17)  # 左眉最左侧
        left_brow_right = landmarks.part(26)  # 右眉最右侧

        # 计算眉毛最外侧点的距离
        brow_width = left_brow_right.x - left_brow_left.x

        # 获取剩余52个点的最小和最大横纵坐标
        min_x = min(point.x for point in landmarks.parts())
        min_y = min(point.y for point in landmarks.parts())
        max_x = max(point.x for point in landmarks.parts())
        max_y = max(point.y for point in landmarks.parts())

        # 根据眉毛宽度调整矩形框的宽度
        w = brow_width
        h = max_y - min_y

        # 获取矩形框的坐标，并根据宽度调整
        x_center = (min_x + max_x) // 2
        y_center = (min_y + max_y) // 2

        # 计算新的矩形框坐标，增加高度（向上加10像素）
        x1 = max(x_center - w // 2, 0)
        y1 = max(y_center - h // 2 - 15, 0)  # 向上加10像素
        x2 = x1 + w
        y2 = y1 + h + 10  # 增加高度

        # 切割图像
        cropped_face = image[y1:y2, x1:x2]

        # 保存剪裁后的图片
        cv2.imwrite(output_path, cropped_face)
        print(f"剪裁后的图片已保存为 {output_path}")
    else:
        print(f"未检测到人脸：{image_path}")


def copy_images_to_new_path(img_paths, dest_dir):
    """
    将多个图片复制到新的目录中
    :param img_paths: 图片路径列表
    :param dest_dir: 目标目录路径
    """
    # 如果目标目录不存在，先创建
    os.makedirs(dest_dir, exist_ok=True)

    for img_path in img_paths:
        if os.path.exists(img_path):
            # 取出文件名
            file_name = os.path.basename(img_path)
            # 拼接目标路径
            dest_path = os.path.join(dest_dir, file_name)
            # 复制文件
            shutil.copy(img_path, dest_path)
            print(f"已复制: {img_path} → {dest_path}")
        else:
            print(f"文件不存在: {img_path}")


# 使用示例
base_path = r'E:\Xyj_magnify\Xyj_dataset_flownet\DFME_Crop_dlib\DFME_Crop\DFME_train'  # 你希望创建文件夹的基础路径

# 加载工作簿
workbook = load_workbook(filename=r'E:\Xyj_magnify\Xyj_dataset_flownet\DFME\training_full_with_AU.xlsx')

# 选择工作表
sheet = workbook['Sheet1']

# 读取整个工作表数据
xlsx_values = []
for row in sheet.iter_rows(min_row=1, max_col=14, values_only=True):
    xlsx_values.append(row)

del xlsx_values[0]  # 删除标题行

# 情感列表
main_path = r"E:\Xyj_magnify\Xyj_dataset_flownet\DFME\DFME_train"


# 遍历 Excel 数据并复制图片
for row in xlsx_values:
    Subject = row[0]
    Filename = row[1]
    OnsetFrame = str(row[2])
    OnsetFrame = OnsetFrame.zfill(5)

    ApexFrame = str(row[3])
    ApexFrame = ApexFrame.zfill(5)
    EmotionData = row[5]

    folder_path = os.path.join(base_path, f"{Filename}")

    # 如果文件夹不存在，则创建它
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # 然后生成目标文件路径
    frame1_revise_path = os.path.join(main_path, f"{Filename}", f"{OnsetFrame}.png")
    frame2_revise_path = os.path.join(main_path, f"{Filename}", f"{ApexFrame}.png")
    frame1_dst_image = os.path.join(folder_path, f"{OnsetFrame}.png")
    frame2_dst_image = os.path.join(folder_path, f"{ApexFrame}.png")

    crop_face(frame1_revise_path, frame1_dst_image)
    crop_face(frame2_revise_path, frame2_dst_image)
