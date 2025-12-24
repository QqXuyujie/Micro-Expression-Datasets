import os
import sys
import cv2
import torch
import numpy as np
from config import Config
from magnet import MagNet
from data import get_gen_ABC, unit_postprocessing, numpy2cuda, resize2d
from callbacks import gen_state_dict
from openpyxl import load_workbook
from os.path import *
import codecs

# config
config = Config()
# Load weights
ep = ''
# weights_file = sorted(
#     [p for p in os.listdir(config.save_dir) if '_loss' in p and '_epoch{}'.format(ep) in p and 'D' not in p],
#     key=lambda x: float(x.rstrip('.pth').split('_loss')[-1])
# )[0]
weights_file = 'magnet_epoch12_loss7.28e-02.pth'
weights_path = os.path.join(weights_file)
ep = int(weights_path.split('epoch')[-1].split('_')[0])
state_dict = gen_state_dict(weights_path)

if torch.cuda.is_available():
    model_test = MagNet().cuda()
else:
    model_test = MagNet().cpu()

model_test.load_state_dict(state_dict)
model_test.eval()
print("Loading weights:", weights_file)

if len(sys.argv) == 1:
    testset = 'myself'
else:
    testset = sys.argv[-1]
testset = testset.split('-')

# 使用示例
base_path = r'E:\Xyj_magnify\Xyj_dataset_flownet\DFME_Crop_dlib\DFME_Crop_Magnify\DFME_testb'  # 你希望创建文件夹的基础路径

workbook = load_workbook(filename=r'E:\Xyj_magnify\Xyj_dataset_flownet\DFME\testB_full_with_AU.xlsx')
sheet = workbook['Sheet1']

# 读取整个工作表数据
xlsx_values = []
for row in sheet.iter_rows(min_row=1, max_col=8, values_only=True):
    xlsx_values.append(row)

del xlsx_values[0]  # 删除标题行

# 情感列表
main_path = r"E:\Xyj_magnify\Xyj_dataset_flownet\DFME_Crop_dlib\DFME_Crop\DFME_testb"
num = 0

# 遍历 Excel 数据并复制图片
for row in xlsx_values:
    Subject = row[0]
    Filename = row[1]
    OnsetFrame = str(row[2])
    OnsetFrame = OnsetFrame.zfill(5)

    ApexFrame = str(row[3])
    ApexFrame = ApexFrame.zfill(5)
    EmotionData = row[5]
    one = os.path.join(main_path, f"{Filename}", f"{OnsetFrame}.png")
    apex = os.path.join(main_path, f"{Filename}", f"{ApexFrame}.png")
    data_loader = get_gen_ABC(config, one, apex, mode='test_on_myself')  # 有的形参有默认值，有的形参没有默认值，那么有默认值的形参要放在没有默认值的形参的后面
    print('Number of test image couples:', data_loader.data_len)
    print(data_loader.paths[0])
    vid_size = (375, 324)

    dst_apex_image = os.path.join(base_path, f"{Filename}_{EmotionData}.jpg")

    # Test
    for amp in [3]:  # [3,5]噪音相对小
        frames = []
        data_loader = get_gen_ABC(config, one, apex, mode='test_on_myself')
        for idx_load in range(0, data_loader.data_len, data_loader.batch_size):
            if (idx_load + 1) % 100 == 0:
                print('{}'.format(idx_load + 1), end=', ')
            batch_A, batch_B = data_loader.gen_test()
            amp_factor = numpy2cuda(amp)
            for _ in range(len(batch_A.shape) - len(amp_factor.shape)):
                amp_factor = amp_factor.unsqueeze(-1)
            with torch.no_grad():
                y_hats = model_test(batch_A, batch_B, 0, 0, amp_factor, mode='evaluate')
            for y_hat in y_hats:
                y_hat = unit_postprocessing(y_hat, vid_size=vid_size)
                frames.append(y_hat)
                if len(frames) >= data_loader.data_len:
                    break
            if len(frames) >= data_loader.data_len:
                break
            if idx_load == 0:
                first_frame = frames
            else:
                apex_frame = frames
        data_loader = get_gen_ABC(config, one, apex, mode='test_on_myself')

        frames = [unit_postprocessing(data_loader.gen_test()[0], vid_size=vid_size)] + frames

        for frame in frames:
            frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
            cv2.imwrite(dst_apex_image, frame)
        print(f"Saved flow image: {apex} to {dst_apex_image}")
        num += 1
        print(num)
