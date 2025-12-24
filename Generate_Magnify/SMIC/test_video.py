import os
import sys
import cv2
import torch
import numpy as np
from config import Config
from magnet import MagNet
from data import get_gen_ABC, unit_postprocessing, numpy2cuda, resize2d
from callbacks import gen_state_dict
from openpyxl import load_workbook
from os.path import *
import codecs


def read_frame_with_padding(main_path, Subject, Filename, OnsetFrame, ApexFrame):
    # 构建文件路径
    frame1_path = os.path.join(main_path, Subject, Filename, f"{Subject}_{OnsetFrame}.jpg")
    frame2_path = os.path.join(main_path, Subject, Filename, f"{Subject}_{ApexFrame}.jpg")

    # 尝试读取 frame1，并根据需要加前导零
    for i in range(4):  # 尝试最多加三次零（i从0到3）
        if os.path.exists(frame1_path):  # 判断路径是否有效
            frame1 = cv2.imread(frame1_path)
            if frame1 is not None:  # 如果读取成功，则跳出循环
                print(f"成功读取 {frame1_path}")
                break
        # 如果路径无效，则在 OnsetFrame 前面加一个0，最多加3次
        OnsetFrame = '0' + str(OnsetFrame)
        frame1_path = os.path.join(main_path, Subject, Filename, f"{Subject}_{OnsetFrame}.jpg")

    # 尝试读取 frame2
    for i in range(4):  # 同样尝试最多加三次零
        if os.path.exists(frame2_path):  # 判断路径是否有效
            frame2 = cv2.imread(frame2_path)
            if frame2 is not None:  # 如果读取成功，则跳出循环
                print(f"成功读取 {frame2_path}")
                break
        # 如果路径无效，则在 ApexFrame 前面加一个0，最多加3次
        ApexFrame = '0' + str(ApexFrame)
        frame2_path = os.path.join(main_path, Subject, Filename, f"{Subject}_{ApexFrame}.jpg")

    return frame1_path, frame2_path


# config
config = Config()
# Load weights
ep = ''
# weights_file = sorted(
#     [p for p in os.listdir(config.save_dir) if '_loss' in p and '_epoch{}'.format(ep) in p and 'D' not in p],
#     key=lambda x: float(x.rstrip('.pth').split('_loss')[-1])
# )[0]
weights_file = 'magnet_epoch12_loss7.28e-02.pth'
weights_path = os.path.join(weights_file)
ep = int(weights_path.split('epoch')[-1].split('_')[0])
state_dict = gen_state_dict(weights_path)

if torch.cuda.is_available():
    model_test = MagNet().cuda()
else:
    model_test = MagNet().cpu()

model_test.load_state_dict(state_dict)
model_test.eval()
print("Loading weights:", weights_file)

if len(sys.argv) == 1:
    testset = 'myself'
else:
    testset = sys.argv[-1]
testset = testset.split('-')

# 使用示例
base_path = r'E:\Xyj_magnify\Xyj_data_magnify\SMIC\SMIC_Magnify\SMIC_Magnify_5_amplify'  # 你希望创建文件夹的基础路径

# 加载工作簿
workbook = load_workbook(filename=r'E:\Xyj_magnify\Xyj_data_magnify\SMIC\smic_apex.xlsx')

# 选择工作表
sheet = workbook['Sheet1']

# 读取整个工作表数据
xlsx_values = []
for row in sheet.iter_rows(min_row=1, max_col=10, values_only=True):
    xlsx_values.append(row)

del xlsx_values[0]  # 删除标题行

# 情感列表
emotions = ['positive', 'negative', 'surprise']
emotion_dict = {0: 'negative', 1: 'positive', 2: 'surprise'}
main_path = r"D:\BaiduNetdiskDownload\SMIC\SMIC_HS_classes"
num = 0

# 遍历 Excel 数据并复制图片
for row in xlsx_values:
    Subject = row[1]
    Filename = row[2]
    Label = row[3]
    OnsetFrame = row[4]
    ApexFrame = row[5]

    # 处理 OnsetFrame
    if OnsetFrame < 100000:
        OnsetFrame = str(OnsetFrame).zfill(6)
    # 处理 ApexFrame
    if ApexFrame < 100000:
        ApexFrame = str(ApexFrame).zfill(6)
    EstimatedEmotion = emotion_dict[int(Label)]

    # 构建源文件路径
    one = os.path.join(main_path, f"{EstimatedEmotion}", Filename, f"reg_image{OnsetFrame}.bmp")
    apex = os.path.join(main_path, f"{EstimatedEmotion}", Filename, f"reg_image{ApexFrame}.bmp")
    data_loader = get_gen_ABC(config, one, apex, mode='test_on_myself')  # 有的形参有默认值，有的形参没有默认值，那么有默认值的形参要放在没有默认值的形参的后面
    print('Number of test image couples:', data_loader.data_len)
    print(data_loader.paths[0])
    vid_size = (310, 340)

    if EstimatedEmotion in emotions:
        dst_folder = os.path.join(base_path, f"sub{Subject}", EstimatedEmotion)
        # 检查目标文件夹是否存在
        if not os.path.exists(dst_folder):
            os.makedirs(dst_folder)
        dst_image = os.path.join(dst_folder, f"{str(Subject).lstrip('0')}_{OnsetFrame}_{ApexFrame}.jpg")
        # Test
        for amp in [5]:  # [3,5]噪音相对小
            frames = []
            data_loader = get_gen_ABC(config, one, apex, mode='test_on_myself')
            for idx_load in range(0, data_loader.data_len, data_loader.batch_size):
                if (idx_load + 1) % 100 == 0:
                    print('{}'.format(idx_load + 1), end=', ')
                batch_A, batch_B = data_loader.gen_test()
                amp_factor = numpy2cuda(amp)
                for _ in range(len(batch_A.shape) - len(amp_factor.shape)):
                    amp_factor = amp_factor.unsqueeze(-1)
                with torch.no_grad():
                    y_hats = model_test(batch_A, batch_B, 0, 0, amp_factor, mode='evaluate')
                for y_hat in y_hats:
                    y_hat = unit_postprocessing(y_hat, vid_size=vid_size)
                    frames.append(y_hat)
                    if len(frames) >= data_loader.data_len:
                        break
                if len(frames) >= data_loader.data_len:
                    break
                if idx_load == 0:
                    first_frame = frames
                else:
                    apex_frame = frames
            data_loader = get_gen_ABC(config, one, apex, mode='test_on_myself')

            frames = [unit_postprocessing(data_loader.gen_test()[0], vid_size=vid_size)] + frames

            for frame in frames:
                frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
                cv2.imwrite(dst_image, frame)
            print(f"Saved flow image: {apex} to {dst_image}")
            num += 1
            print(num)
